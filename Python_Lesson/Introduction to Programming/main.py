import openpyxl

path = "C:\\Users\\Ibrahim S Saliu\\OneDrive\\Documents\\UPSKILLING\\Software development\\awsCloudProjectBootcamp\\awsCloudProjectBootcamp\\Python_Lesson\\Introduction to Programming\\inventory.xlsx"
inventory_file = openpyxl.load_workbook(path)
product_list = inventory_file.active

products_per_supplier = {}
total_value_per_supplier = {}
for product_row in range(2,
                         product_list.max_row + 1):  # range creates a list of number [0,1,2,3,4,5........], We are starting from line 2
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    print(supplier_name)

    # to calculate the number of products we have per supplier and list the supplier with the number of products

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a new supplier.... ")
        products_per_supplier[supplier_name] = 1

    # to calculate the total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + (inventory * price)

    else:
        total_value_per_supplier[supplier_name] = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)

